"""
Firewall Rule Checker - Analysis Results Formatter
Formats and consolidates the firewall analysis results.

Author: Vishal Patil
Email: vp26781@gmail.com
"""

import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from openpyxl import load_workbook

class AnalysisResultsFormatter:
    """Handles formatting and consolidation of firewall analysis results."""
    
    def __init__(self):
        self.setup_logging()
        self.base_dir = Path(__file__).parent
        self.analysis_file = self.base_dir / "output_external_internal.xlsx"

    def setup_logging(self) -> None:
        """Configure logging with both file and console output."""
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f"analysis_formatter_{timestamp}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - [%(name)s] %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger("AnalysisResultsFormatter")

    def load_analysis_results(self) -> Optional[pd.DataFrame]:
        """Load analysis results from Excel file with validation."""
        try:
            self.logger.info(f"Loading analysis results from {self.analysis_file}")
            
            if not self.analysis_file.exists():
                raise FileNotFoundError(f"Analysis file not found: {self.analysis_file}")
            
            df = pd.read_excel(self.analysis_file)
            
            # Validate required columns
            required_columns = ['Type', 'Excel Row']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error loading analysis results: {str(e)}")
            return None

    def consolidate_results(self, df: pd.DataFrame) -> pd.DataFrame:
        """Consolidate analysis results by combining findings for the same rule."""
        self.logger.info("Consolidating analysis results...")
        
        try:
            # Create a mapping of row numbers to findings
            row_findings: Dict[int, List[str]] = {}
            row_data: Dict[int, pd.Series] = {}
            
            for _, row in df.iterrows():
                excel_row = row['Excel Row']
                finding_type = row['Type']
                
                if pd.isna(excel_row):
                    continue
                    
                if excel_row not in row_findings:
                    row_findings[excel_row] = []
                    row_data[excel_row] = row
                    
                row_findings[excel_row].append(finding_type)
            
            # Create consolidated findings
            consolidated_data = []
            for row_num, findings in row_findings.items():
                base_row = row_data[row_num]
                
                # Combine findings
                findings_str = " - ".join(sorted(set(findings)))
                
                # Create consolidated row
                new_row = base_row.copy()
                new_row['Type'] = findings_str
                consolidated_data.append(new_row)
            
            # Create new DataFrame
            consolidated_df = pd.DataFrame(consolidated_data)
            
            # Sort by Excel Row
            consolidated_df = consolidated_df.sort_values('Excel Row')
            
            self.logger.info(f"Consolidated {len(df)} findings into {len(consolidated_df)} unique rules")
            return consolidated_df
            
        except Exception as e:
            self.logger.error(f"Error consolidating results: {str(e)}")
            raise

    def format_results(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply formatting to the analysis results DataFrame."""
        try:
            self.logger.info("Formatting analysis results...")
            
            # Add categorization for findings
            df['Communication Type'] = df['Type'].apply(self.categorize_communication)
            
            # Reorder columns
            column_order = [
                'Excel Row',
                'Rule Number',
                'Type',
                'Communication Type',
                'Source',
                'Destination',
                'Service',
                'Public IPs',
                'Private IPs'
            ]
            
            # Keep only available columns
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error formatting results: {str(e)}")
            raise

    def categorize_communication(self, rule_type: str) -> str:
        """Categorize the type of communication."""
        if 'Public Source' in rule_type:
            return 'Inbound from Internet'
        elif 'Public Destination' in rule_type:
            return 'Outbound to Internet'
        else:
            return 'Internal Communication'

    def save_results(self, original_df: pd.DataFrame, formatted_df: pd.DataFrame) -> None:
        """Save results to Excel with original and formatted sheets."""
        try:
            self.logger.info(f"Saving results to {self.analysis_file}")
            
            # Save with original and formatted sheets
            with pd.ExcelWriter(self.analysis_file, engine='openpyxl', mode='a') as writer:
                # Remove the 'Formatted_Results' sheet if it exists
                if 'Formatted_Results' in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index('Formatted_Results')
                    writer.book.remove(writer.book.worksheets[idx])
                
                # Write the formatted findings to a new sheet
                formatted_df.to_excel(writer, sheet_name='Formatted_Results', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Formatted_Results']
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            self.logger.info("Results saved successfully")
            
        except Exception as e:
            self.logger.error(f"Error saving results: {str(e)}")
            raise

    def process_results(self) -> bool:
        """Main processing function."""
        try:
            # Load results
            original_df = self.load_analysis_results()
            if original_df is None:
                return False
            
            # Process results
            formatted_df = self.consolidate_results(original_df.copy())
            formatted_df = self.format_results(formatted_df)
            
            # Save results
            self.save_results(original_df, formatted_df)
            
            self.logger.info("Results processing completed successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error processing results: {str(e)}")
            return False

def main():
    """Main execution function."""
    try:
        formatter = AnalysisResultsFormatter()
        
        if formatter.process_results():
            print("Analysis results formatting completed successfully.")
            return 0
        else:
            print("Analysis results formatting failed. Check logs for details.")
            return 1
            
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        return 1
    except Exception as e:
        print(f"Critical error: {str(e)}")
        return 1

if __name__ == "__main__":
    exit(main())
