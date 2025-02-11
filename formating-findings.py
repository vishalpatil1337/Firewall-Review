"""
Firewall Rule Checker - Findings Formatter
Formats and consolidates the analysis findings.

Author: Vishal Patil
Email: vp26781@gmail.com
"""

import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from openpyxl import load_workbook

class FindingsFormatter:
    """Handles formatting and consolidation of firewall analysis findings."""
    
    def __init__(self):
        self.setup_logging()
        self.base_dir = Path(__file__).parent
        self.findings_file = self.base_dir / "output_cde-oos-findings.xlsx"

    def setup_logging(self) -> None:
        """Configure logging with both file and console output."""
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f"findings_formatter_{timestamp}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - [%(name)s] %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger("FindingsFormatter")

    def load_findings(self) -> Optional[pd.DataFrame]:
        """Load findings from Excel file with validation."""
        try:
            self.logger.info(f"Loading findings from {self.findings_file}")
            
            if not self.findings_file.exists():
                raise FileNotFoundError(f"Findings file not found: {self.findings_file}")
            
            df = pd.read_excel(self.findings_file)
            
            # Validate required columns
            required_columns = ['Type', 'Excel Row']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error loading findings: {str(e)}")
            return None

    def consolidate_findings(self, df: pd.DataFrame) -> pd.DataFrame:
        """Consolidate findings by combining duplicate rows."""
        self.logger.info("Consolidating findings...")
        
        try:
            # Create a mapping of Excel rows to findings
            row_findings: Dict[int, List[str]] = {}
            row_data: Dict[int, pd.Series] = {}
            
            for _, row in df.iterrows():
                excel_row = row['Excel Row']
                finding_type = row['Type']
                
                if pd.isna(excel_row):
                    continue
                    
                if excel_row not in row_findings:
                    row_findings[excel_row] = []
                    row_data[excel_row] = row
                    
                row_findings[excel_row].append(finding_type)
            
            # Create consolidated findings
            consolidated_data = []
            for row_num, findings in row_findings.items():
                base_row = row_data[row_num]
                
                # Combine findings
                findings_str = " - ".join(sorted(set(findings)))
                
                # Create consolidated row
                new_row = base_row.copy()
                new_row['Type'] = findings_str
                consolidated_data.append(new_row)
            
            # Create new DataFrame
            consolidated_df = pd.DataFrame(consolidated_data)
            
            # Sort by Excel Row
            consolidated_df = consolidated_df.sort_values('Excel Row')
            
            self.logger.info(f"Consolidated {len(df)} findings into {len(consolidated_df)} unique rules")
            return consolidated_df
            
        except Exception as e:
            self.logger.error(f"Error consolidating findings: {str(e)}")
            raise

    def format_findings(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply formatting to the findings DataFrame."""
        try:
            self.logger.info("Formatting findings...")
            
            # Reorder columns
            column_order = [
                'Excel Row',
                'Type',
                'Source',
                'Destination',
                'Service',
                'Matching_Source_Ranges',
                'Matching_Destination_Ranges'
            ]
            
            # Keep only available columns
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error formatting findings: {str(e)}")
            raise

    def save_findings(self, original_df: pd.DataFrame, formatted_df: pd.DataFrame) -> None:
        """Save findings to Excel with original and formatted sheets."""
        try:
            self.logger.info(f"Saving findings to {self.findings_file}")
            
            # First, save the original data to maintain the original formatting
            with pd.ExcelWriter(self.findings_file, engine='openpyxl', mode='a') as writer:
                # Remove the 'Formatted_Findings' sheet if it exists
                if 'Formatted_Findings' in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index('Formatted_Findings')
                    writer.book.remove(writer.book.worksheets[idx])
                
                # Write the formatted findings to a new sheet
                formatted_df.to_excel(writer, sheet_name='Formatted_Findings', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Formatted_Findings']
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            self.logger.info("Findings saved successfully")
            
        except Exception as e:
            self.logger.error(f"Error saving findings: {str(e)}")
            raise

    def process_findings(self) -> bool:
        """Main processing function."""
        try:
            # Load findings
            original_df = self.load_findings()
            if original_df is None:
                return False
            
            # Process findings
            formatted_df = self.consolidate_findings(original_df.copy())
            formatted_df = self.format_findings(formatted_df)
            
            # Save results
            self.save_findings(original_df, formatted_df)
            
            self.logger.info("Findings processing completed successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error processing findings: {str(e)}")
            return False

def main():
    """Main execution function."""
    try:
        formatter = FindingsFormatter()
        
        if formatter.process_findings():
            print("Findings formatting completed successfully.")
            return 0
        else:
            print("Findings formatting failed. Check logs for details.")
            return 1
            
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        return 1
    except Exception as e:
        print(f"Critical error: {str(e)}")
        return 1

if __name__ == "__main__":
    exit(main())
